#file management  ---done---
#data management ---done---
#sheduling and remainders  
import os 
import shutil
#open listes of folders and then specific is word ppt or .txt files and organize it in specific files 
def filemanager(source):
    txt = 'C:/Users/hcc/Downloads/codealpha_tasks--main/txt'
    word='C:/Users/hcc/Downloads/codealpha_tasks--main/doc'
    pptt='C:/Users/hcc/Downloads/codealpha_tasks--main/ppt'
    os.makedirs(txt, exist_ok=True)
    os.makedirs(word, exist_ok=True)
    os.makedirs(pptt, exist_ok=True)
#path exist or no 
    if  not os.path.exists(source):
      print('the source file not exist')
      return  
    #file exiist so we an read it 
    for folder_name in os.listdir(source):
     folder_path = os.path.join(source, folder_name)
     if os.path.isdir(folder_path):
        
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
       
            
            if os.path.isfile(file_path):
               if filename.endswith('.txt'):
                        shutil.copy(file_path, txt)
               elif filename.endswith('.ppt') or filename.endswith('.pptm'):
                        shutil.copy(file_path, pptt)
               elif filename.endswith('.doc') or filename.endswith('.docx'):
                        shutil.copy(file_path, word)
            else:

             print(" the file not supported:",file_path)
    print('files moved successfully')
    
 
#2nd def manger data entry for employment 
import openpyxl
from openpyxl import Workbook



def data_entry(Firstname , Lastname , email,phone,adress):
   file = 'data.xlsx'
    
   if os.path.exists(file):
        # Load the existing workbook
        wb = openpyxl.load_workbook(file)
        sheet = wb['Sheet1']
   else: 
      # Create a new workbook if it doesn't exist
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Sheet1'
        # Add headers to the first row if creating a new workbook
        sheet.append(['First Name', 'Last Name', 'Email', 'Phone', 'Address'])
        wb.save(file)
    
    # Reload the sheet in case it was modified during the session
   wb = openpyxl.load_workbook(file)
   sheet = wb['Sheet1']
    
    # Check for duplicates
   for row in sheet.iter_rows(min_row=2, values_only=True):
        if (row[0].strip().lower() == Firstname.strip().lower() and 
            row[1].strip().lower() == Lastname.strip().lower() and 
            row[2].strip().lower() == email.strip().lower() and 
            row[3].strip().lower() == phone.strip().lower()):
            print(f'Entry for {Firstname} {Lastname} already exists.')
            return
    
    # Add the new entry if no duplicate is found
   sheet.append([Firstname, Lastname, email, phone, adress])
   wb.save(file)
   print(f'Entry for {Firstname} {Lastname} added successfully.')
    
   wb.save(file)
def data_extract(Firstname,Lastname):
    file = 'data.xlsx'
    if os.path.exists(file):
        # Load the existing workbook
        wb = openpyxl.load_workbook(file)
        sheet = wb['Sheet1']
    else:
        print('The file does not exist.')
        return
    
    # Search for the specified employee
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0].strip().lower() == Firstname.strip().lower() and row[1].strip().lower() == Lastname.strip().lower():
            print('Employee found:', row[2:])  # Print the employee's details (email, phone, address)
            wb.close()
            return
    
    print('Employee not found.')
    wb.close()

def data_cleaning(Firstname , Lastname , email,phone,adress):
    file='data.xlsx'
    if os.path.exists(file):
        wb = openpyxl.load_workbook(file)
        sheet = wb['Sheet1']
    else:
        print('FILE NOT FOUND ')
    #REMOVE DATA OF EMPLOYEE FROM WORKBOOK
    for row in sheet.iter_rows(min_row=2):
        if (row[0].value.strip().lower() == Firstname.strip().lower() and 
            row[1].value.strip().lower() == Lastname.strip().lower() and 
            row[2].value.strip().lower() == email.strip().lower() and 
            row[3].value.strip() == phone.strip() and 
            row[4].value.strip().lower() == adress.strip().lower()):
            # Delete the row
            sheet.delete_rows(row[0].row)
            print('Employee successfully removed.')
            wb.save(file)  # Save the workbook after deletion
            wb.close()
            return
    
    print('Employee not found.')
    wb.close()
def salray_calculate(nb_projects, work_weekend, Firstname, Lastname):
    base_salary = 500
    # Each project adds $5
    totale = base_salary + nb_projects * 5
    if work_weekend:
        totale += totale * 0.20
    
    file = 'data.xlsx'
    if os.path.exists(file):
        wb = openpyxl.load_workbook(file)
        sheet = wb['Sheet1']
    else:
        print('FILE NOT FOUND')
        return
    if sheet.max_column < 6:
       
        # Adding header for the sixth column
        sheet.cell(row=1, column=6, value='Salary')
    for row in sheet.iter_rows(min_row=2):
        first_name_cell = row[0].value
        last_name_cell = row[1].value

        if (first_name_cell and last_name_cell and
            first_name_cell.strip().lower() == Firstname.strip().lower() and
            last_name_cell.strip().lower() == Lastname.strip().lower()):

           row[5].value = totale

            
        wb.save(file)
        print(f'Salary updated successfully for {Firstname} {Lastname}.')
        return
    
    print(f'Employee {Firstname} {Lastname} not found.')


# Example usage:
#salray_calculate(30, True, 'reem', 'zouhbu')

#automation task 

source = r'C:\Users\hcc\Downloads\codealpha_tasks--main\source'
#filemanager(source)
data_entry('reem','zouhby','reemz@gmail.com','99999','halba ')
data_entry('frah','zouhbi','fff@gmail.com','9789','halba')
data_entry('srour', 'hammoud', 's@gmail.com', '99', 'Tripoli')
data_entry('majdeline', 'allawa', 'majdelinez@gmail.com', '88888', 'Tripoli')
#data_extract('reem','zouhby')
#data_cleaning('majdeline', 'allawa', 'majdelinez@gmail.com', '88888', 'Tripoli')
salray_calculate(30,True,'reem','zouhby')
